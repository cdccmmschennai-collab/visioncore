"""Writer for the **AI Output** workbook.

Reproduces the supplied reference file cell-for-cell: one sheet named
`<TAG> - <DESCRIPTION>`, a merged title and source line, a three-column table
(Field / Extracted Value / Data Quality) in the fixed FIELDS order, a Remarks
row with no quality mark, and the two-row legend beneath.

Palette and metrics were read off the reference workbook, not invented:
  header fill #14345A, header text #FFFFFF
  field-label fill #EEF2F7, label text #1F3864
  Confirmed  fill #E2EFDA, text #375623
  Verify     fill #FFC000, text #7F4E00
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.fields import (
    FIELDS,
    QUALITY_CONFIRMED,
    REMARKS_FIELD,
    is_blank,
    quality_of,
    value_of,
)

TITLE_COLOR = "FF1F3864"
SUBTITLE_COLOR = "FF595959"
HEADER_FILL = "FF14345A"
LABEL_FILL = "FFEEF2F7"
CONFIRMED_FILL = "FFE2EFDA"
CONFIRMED_TEXT = "FF375623"
VERIFY_FILL = "FFFFC000"
VERIFY_TEXT = "FF7F4E00"
# Not a quality mark — the AI found nothing to grade, so it gets a neutral
# gray rather than the amber "Verify" would wrongly imply a reading exists.
MISSING_LABEL = "Data not available"
MISSING_FILL = "FFE0E0E0"
MISSING_TEXT = "FF595959"

_THIN = Side(style="thin", color="FFB4C6E7")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Rows that need extra height because their content wraps.
TALL_ROWS = {"hazardous_classification": 33.0, "additional_information": 48.0}


def _sheet_title(tag_number: str, description: str) -> str:
    """Excel caps sheet names at 31 chars and forbids : \\ / ? * [ ]."""
    title = f"{tag_number} - {description}"
    for bad in ':\\/?*[]':
        title = title.replace(bad, "-")
    return title[:31]


def build_ai_workbook(payload: dict, tag_number: str, description: str,
                      source_filenames: list[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(tag_number, description)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 27.0
    ws.column_dimensions["B"].width = 58.0
    ws.column_dimensions["C"].width = 16.0
    ws.column_dimensions["D"].width = 3.0

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    ws["A1"] = f"Full Technical Data Sheet  \u2014  {description}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=TITLE_COLOR)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26.1

    ws.merge_cells("A2:C2")
    sources = ", ".join(source_filenames) if source_filenames else "\u2014"
    ws["A2"] = f"Source: Nameplate photo(s) \u2014 {sources}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=SUBTITLE_COLOR)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[2].height = 27.95

    # ── Table header (row 4; row 3 is deliberately blank) ────────────────────
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for col, label in enumerate(("Field", "Extracted Value", "Data Quality"), start=1):
        cell = ws.cell(row=4, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[4].height = 20.1

    label_font = Font(name="Calibri", size=11, bold=True, color=TITLE_COLOR)
    label_fill = PatternFill("solid", fgColor=LABEL_FILL)
    body_font = Font(name="Calibri", size=11)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    centred = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = 5
    for field in FIELDS:
        a = ws.cell(row=row, column=1, value=field.ai_label)
        a.font, a.fill, a.alignment, a.border = label_font, label_fill, left_top, BORDER

        value = value_of(payload, field.key)
        b = ws.cell(row=row, column=2, value=value)
        b.font, b.alignment, b.border = body_font, left_top, BORDER

        missing = is_blank(value)
        quality = quality_of(payload, field.key)
        confirmed = quality == QUALITY_CONFIRMED
        c = ws.cell(row=row, column=3, value=MISSING_LABEL if missing else quality)
        c.font = Font(
            name="Calibri", size=11, bold=True,
            color=MISSING_TEXT if missing else (CONFIRMED_TEXT if confirmed else VERIFY_TEXT),
        )
        c.fill = PatternFill(
            "solid",
            fgColor=MISSING_FILL if missing else (CONFIRMED_FILL if confirmed else VERIFY_FILL),
        )
        c.alignment, c.border = centred, BORDER

        ws.row_dimensions[row].height = TALL_ROWS.get(field.key, 18.0)
        row += 1

    # ── Remarks: a finding, not a data field, so it carries no quality mark ──
    a = ws.cell(row=row, column=1, value=REMARKS_FIELD.ai_label)
    a.font, a.fill, a.alignment, a.border = label_font, label_fill, left_top, BORDER
    b = ws.cell(row=row, column=2, value=str(payload.get("remarks", "") or ""))
    b.font, b.alignment, b.border = body_font, left_top, BORDER
    ws.row_dimensions[row].height = 48.0
    remarks_row = row

    # ── Legend, one blank row below the table ────────────────────────────────
    legend_row = remarks_row + 2
    ws.cell(row=legend_row, column=1, value="Data Quality Legend").font = Font(
        name="Calibri", size=11, bold=True, color=TITLE_COLOR
    )

    legend = (
        (QUALITY_CONFIRMED, CONFIRMED_FILL, CONFIRMED_TEXT,
         "Clearly legible / directly read from the nameplate"),
        ("Verify", VERIFY_FILL, VERIFY_TEXT,
         "Inferred, referenced, worn/partly legible, or not printed on the tag "
         "\u2014 recommend field verification"),
        (MISSING_LABEL, MISSING_FILL, MISSING_TEXT,
         "The AI found no such detail on the nameplate \u2014 nothing to confirm or verify"),
    )
    for offset, (label, fill, text_color, note) in enumerate(legend, start=1):
        r = legend_row + offset
        key_cell = ws.cell(row=r, column=1, value=label)
        key_cell.font = Font(name="Calibri", size=11, bold=True, color=text_color)
        key_cell.fill = PatternFill("solid", fgColor=fill)
        key_cell.alignment = Alignment(horizontal="center")
        key_cell.border = BORDER

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        note_cell = ws.cell(row=r, column=2, value=note)
        note_cell.font = Font(name="Calibri", size=10, color=SUBTITLE_COLOR)
        note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.print_area = f"A1:C{legend_row + len(legend)}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    for col in range(1, 4):
        get_column_letter(col)   # widths already set above

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
