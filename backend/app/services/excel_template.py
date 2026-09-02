"""Writer for the **Template Output** workbook.

A single `Asset Tag` sheet with 17 columns in fixed order (S.NO → INPUT
PHOTO), yellow Arial-10 header, frozen header row, and an autofilter across
the used range. The last two columns, AI OUTPUT EXCEL and INPUT PHOTO, are
clickable hyperlinks — driven off `record["ai_excel_url"]` /
`record["input_photo_url"]`, signed links built from the tag's own
tag_number by `app.services.download_links`, never off the row's other data,
so one tag's links can never point at another tag's files. A signed link
resolves the actual file at click time over HTTP, so it opens correctly
wherever the exported workbook is later opened — unlike a local filesystem
path, which is only ever valid on the machine that generated the file. A tag
with nothing to link to gets "Not available" text instead of a dead link.

Two conventions were read off the reference row and are reproduced faithfully:

* **Amber fill (#FFC000)** on a value cell whose AI quality mark is `Verify` —
  in the reference, EQPT HAZARDOUS CLASSIFICATION carried this.
* **Blue font (#0070C0)** on a value a reviewer supplied or corrected, where the
  AI had not read it cleanly — in the reference, MAKE (`JC VALVES`) carried
  this, and the REMARKS cell explained that only a logo was visible.

Blank means blank: where the AI reported `Not present on nameplate`, the
Template cell is left empty, exactly as the reference does for MODEL, PART NO,
WEIGHT and COUNTRY.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.services.fields import (
    FIELDS,
    QUALITY_VERIFY,
    REMARKS_FIELD,
    is_blank,
    quality_of,
    value_of,
)

HEADER_FILL = PatternFill("solid", fgColor="FFFFFF00")
AMBER_FILL = PatternFill("solid", fgColor="FFFFC000")
REVIEWER_FONT_COLOR = "FF0070C0"
HYPERLINK_FONT_COLOR = "FF0563C1"

ARIAL_HEADER = Font(name="Arial", size=10, bold=True, color="FF000000")
ARIAL_BODY = Font(name="Arial", size=10)
ARIAL_BODY_REVIEWED = Font(name="Arial", size=10, color=REVIEWER_FONT_COLOR)
ARIAL_HYPERLINK = Font(name="Arial", size=10, color=HYPERLINK_FONT_COLOR, underline="single")

AI_OUTPUT_HEADER = "AI OUTPUT EXCEL"
INPUT_PHOTO_HEADER = "INPUT PHOTO"

ASSET_TAG_HEADERS: tuple[str, ...] = (
    "S.NO",
    *(f.template_header for f in FIELDS),
    REMARKS_FIELD.template_header,
    AI_OUTPUT_HEADER,
    INPUT_PHOTO_HEADER,
)

#: Widths lifted from the reference workbook, keyed by header text so they
#: survive any future column reordering in FIELDS.
COLUMN_WIDTHS: dict[str, float] = {
    "S.NO": 5.55, "TAG NUMBER": 17.11, "EQUIPMENT DESCRIPTION": 41.33,
    "SIZE/DIMENSION": 19.0, "MAKE (ASSET)": 88.66, "MODEL": 44.33,
    "SERIAL NO": 37.33, "PART NO": 9.33, "WEIGHT": 8.44, "COUNTRY": 10.44,
    "YEAR OF MANUFACTURE YYYY": 29.33, "MONTH OF MANUFACTURE MM": 29.89,
    "EQPT HAZARDOUS CLASSIFICATION": 104.33, "ADDITIONAL INFORMATION": 255.66,
    "REMARKS": 132.89, AI_OUTPUT_HEADER: 18.0, INPUT_PHOTO_HEADER: 16.0,
}

WRAP_HEADERS = {"EQPT HAZARDOUS CLASSIFICATION", "ADDITIONAL INFORMATION", "REMARKS"}


def _write_asset_tags_header(ws: Worksheet) -> None:
    for col, header in enumerate(ASSET_TAG_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = ARIAL_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = COLUMN_WIDTHS.get(header, 18.0)
    ws.row_dimensions[1].height = 59.25
    ws.freeze_panes = "A2"


def _write_asset_tag_row(ws: Worksheet, row_idx: int, serial: int, record: dict) -> None:
    """`record` carries the final payload (plus the AI's original for diffing)."""
    payload = record["payload"]
    ai_payload = record.get("ai_payload") or {}

    sno = ws.cell(row=row_idx, column=1, value=f"{serial:04d}")
    sno.font = ARIAL_BODY
    sno.alignment = Alignment(horizontal="center", vertical="center")

    col = 2
    for field in FIELDS:
        value = value_of(payload, field.key)
        display = "" if is_blank(value) else value
        cell = ws.cell(row=row_idx, column=col, value=display or None)

        # Force text format so Excel never renders "02" as the number 2.
        if field.key in ("year_of_manufacture", "month_of_manufacture"):
            cell.number_format = "@"

        # Blue where a reviewer supplied or changed the value the AI returned.
        ai_value = value_of(ai_payload, field.key) if ai_payload else value
        reviewer_supplied = bool(display) and ai_value.strip() != value.strip()
        cell.font = ARIAL_BODY_REVIEWED if reviewer_supplied else ARIAL_BODY

        # Amber where the AI flagged the reading as needing field verification.
        if display and quality_of(payload, field.key) == QUALITY_VERIFY:
            cell.fill = AMBER_FILL

        if field.template_header in WRAP_HEADERS:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        col += 1

    remarks = ws.cell(row=row_idx, column=col, value=str(payload.get("remarks", "") or "") or None)
    remarks.font = ARIAL_BODY
    remarks.alignment = Alignment(vertical="top", wrap_text=True)
    col += 1

    _write_hyperlink_cell(ws, row_idx, col, record.get("ai_excel_url"), "View AI Output")
    col += 1
    _write_hyperlink_cell(ws, row_idx, col, record.get("input_photo_url"), "View Photo")

    ws.row_dimensions[row_idx].height = 25.5


def _write_hyperlink_cell(ws: Worksheet, row_idx: int, col: int, url: str | None,
                          link_text: str) -> None:
    """Write a clickable, friendly-text hyperlink, or 'Not available' for a tag
    with no corresponding file (e.g. a tag with no photo attached).

    `url` is a ready-made signed link from app.services.download_links, built
    from the tag's own tag_number — never a raw filesystem path, so it opens
    correctly regardless of which machine later opens this workbook.
    """
    cell = ws.cell(row=row_idx, column=col)
    if url:
        cell.value = link_text
        cell.hyperlink = url
        cell.font = ARIAL_HYPERLINK
    else:
        cell.value = "Not available"
        cell.font = ARIAL_BODY
    cell.alignment = Alignment(horizontal="center", vertical="center")


def build_template_workbook(records: list[dict]) -> bytes:
    """`records` is a list of dicts with keys: payload, ai_payload.
    One row per record — a single tag for a per-tag download, or every tag in a
    batch for the batch-level download.
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "Asset Tag"
    _write_asset_tags_header(ws)
    for offset, record in enumerate(records):
        _write_asset_tag_row(ws, row_idx=2 + offset, serial=offset + 1, record=record)

    last_row = max(2, 1 + len(records))
    ws.auto_filter.ref = f"A1:Q{last_row}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
